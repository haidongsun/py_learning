#!/usr/bin/env python3
"""
MCP Server for Excel File Operations.

Provides comprehensive Excel manipulation tools covering reading, writing,
formatting, sheet management, data analysis, formulas, and import/export.
"""

from __future__ import annotations

import csv
import io
import json
import os
import re
from datetime import datetime
from typing import Any, Optional

import openpyxl
import pandas as pd
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter, column_index_from_string
from pydantic import BaseModel, Field, field_validator, ConfigDict

from mcp.server.fastmcp import FastMCP

mcp = FastMCP("excel_mcp")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_MAX_ROWS_DISPLAY = 200
_MAX_COLS_DISPLAY = 50

# ---------------------------------------------------------------------------
# Enums for options
# ---------------------------------------------------------------------------

BORDER_STYLES = [
    "thin", "medium", "thick", "dashed", "dotted", "double",
    "hair", "mediumDashed", "dashDot", "mediumDashDot",
    "dashDotDot", "mediumDashDotDot", "slantDashDot",
]

HORIZONTAL_ALIGNMENTS = [
    "left", "center", "right", "fill", "justify",
    "centerContinuous", "distributed",
]

VERTICAL_ALIGNMENTS = [
    "top", "center", "bottom", "justify", "distributed",
]


# ---------------------------------------------------------------------------
# Helper utilities
# ---------------------------------------------------------------------------

def _resolve_path(filepath: str) -> str:
    """Normalise and validate file path."""
    path = os.path.abspath(os.path.expanduser(filepath))
    if not os.path.splitext(path)[1].lower() in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        if not os.path.exists(path):
            path += ".xlsx"
    return path


def _safe_open(filepath: str, data_only: bool = False) -> openpyxl.Workbook:
    """Open a workbook with consistent error handling."""
    path = _resolve_path(filepath)
    if not os.path.exists(path):
        raise FileNotFoundError(f"File not found: {path}")
    return openpyxl.load_workbook(path, data_only=data_only)


def _serialise_value(val: Any) -> Any:
    """Convert cell values to JSON-serialisable types."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, (int, float, bool, str)):
        return val
    return str(val)


def _read_sheet_data(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int = 1,
    end_row: int | None = None,
    start_col: int = 1,
    end_col: int | None = None,
    with_headers: bool = True,
) -> dict:
    """Read data from a worksheet and return as structured dict."""
    if end_row is None:
        end_row = ws.max_row
    if end_col is None:
        end_col = ws.max_column

    end_row = min(end_row, ws.max_row)
    end_col = min(end_col, ws.max_column)

    rows = []
    headers = []
    row_start = start_row

    if with_headers:
        headers = [
            str(ws.cell(row=start_row, column=c).value or f"Column{c}")
            for c in range(start_col, end_col + 1)
        ]
        row_start = start_row + 1
    else:
        headers = [get_column_letter(c) for c in range(start_col, end_col + 1)]

    for r in range(row_start, end_row + 1):
        row_data = {}
        for c_idx, c in enumerate(range(start_col, end_col + 1)):
            val = ws.cell(row=r, column=c).value
            row_data[headers[c_idx]] = _serialise_value(val)
        rows.append(row_data)

    return {
        "headers": headers,
        "rows": rows,
        "row_count": len(rows),
        "start_row": start_row,
        "end_row": end_row,
        "start_col": get_column_letter(start_col),
        "end_col": get_column_letter(end_col),
    }


def _format_markdown_table(data: dict, max_rows: int = 50) -> str:
    """Format structured sheet data as a Markdown table."""
    if not data["rows"]:
        return "(empty sheet)"

    headers = data["headers"]
    rows = data["rows"][:max_rows]

    lines = ["| " + " | ".join(headers) + " |"]
    lines.append("|" + "|".join("---" for _ in headers) + "|")
    for row in rows:
        cells = [str(row.get(h, "")) for h in headers]
        lines.append("| " + " | ".join(cells) + " |")

    if len(data["rows"]) > max_rows:
        lines.append(f"\n*... {len(data['rows']) - max_rows} more rows not shown*")

    return "\n".join(lines)


def _build_border(
    style: str = "thin",
    top: bool = True, bottom: bool = True,
    left: bool = True, right: bool = True,
) -> Border:
    """Build an openpyxl Border object from parameters."""
    side = Side(style=style)
    return Border(
        top=side if top else None,
        bottom=side if bottom else None,
        left=side if left else None,
        right=side if right else None,
    )


# ===================================================================
# Pydantic input models
# ===================================================================

class FilePathInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)
    filepath: str = Field(
        ..., description="Path to the Excel file (e.g., 'C:/data/report.xlsx')",
        min_length=1, max_length=4096,
    )


class SheetInput(FilePathInput):
    sheet: str = Field(
        default="Sheet1",
        description="Sheet name (default: 'Sheet1')",
        min_length=1, max_length=31,
    )


class ReadRangeInput(SheetInput):
    start_cell: str = Field(
        default="A1", description="Top-left cell (e.g., 'A1')",
    )
    end_cell: Optional[str] = Field(
        default=None, description="Bottom-right cell (e.g., 'F20'). Auto-detects if omitted.",
    )
    preview_only: bool = Field(
        default=False, description="If True, returns only a Markdown preview (first 50 rows).",
    )

    @field_validator("start_cell", "end_cell")
    @classmethod
    def _valid_cell(cls, v: str | None) -> str | None:
        if v and not re.match(r"^[A-Z]{1,3}\d+$", v.upper()):
            raise ValueError(f"Invalid cell reference: {v}")
        return v.upper() if v else v


class CellWriteInput(SheetInput):
    cell: str = Field(..., description="Cell reference (e.g., 'B3')")
    value: Any = Field(..., description="Value to write (str, int, float, or None)")

    @field_validator("cell")
    @classmethod
    def _valid_cell(cls, v: str) -> str:
        if not re.match(r"^[A-Z]{1,3}\d+$", v.upper()):
            raise ValueError(f"Invalid cell reference: {v}")
        return v.upper()


class RowsWriteInput(SheetInput):
    start_cell: str = Field(default="A1", description="Top-left starting cell")
    data: list[list[Any]] = Field(
        ..., description="2D list of data to write, e.g. [['Name','Age'],['Alice',30]]",
    )
    header: bool = Field(
        default=True, description="Whether the first row contains headers.",
    )


class TabularWriteInput(SheetInput):
    data: list[dict[str, Any]] = Field(
        ..., description="List of dicts, e.g. [{'Name':'Alice','Age':30}]",
    )


class SheetCopyInput(FilePathInput):
    source_sheet: str = Field(..., description="Sheet to copy from")
    target_sheet: str = Field(..., description="Name for the new copied sheet")


class ColWidthInput(SheetInput):
    column: str = Field(..., description="Column letter (e.g., 'A', 'B') or column range (e.g., 'A:C')")
    width: float = Field(..., description="Column width in character units", ge=0, le=255)

    @field_validator("column")
    @classmethod
    def _valid_col(cls, v: str) -> str:
        if not re.match(r"^[A-Z]{1,3}(:[A-Z]{1,3})?$", v.upper()):
            raise ValueError(f"Invalid column reference: {v}")
        return v.upper()


class RowHeightInput(SheetInput):
    row: int = Field(..., description="Row number", ge=1, le=1048576)
    height: float = Field(..., description="Row height in points", ge=0, le=409)


class FormatRangeInput(SheetInput):
    range_str: str = Field(
        ..., description="Cell range (e.g., 'A1:C10' or 'A:A')",
    )
    bold: Optional[bool] = Field(default=None, description="Set font bold")
    font_size: Optional[int] = Field(default=None, description="Font size", ge=1, le=409)
    font_color: Optional[str] = Field(default=None, description="Font hex color (e.g., 'FF0000')")
    bg_color: Optional[str] = Field(default=None, description="Background/fill hex color (e.g., 'FFFF00')")
    horizontal: Optional[str] = Field(
        default=None, description=f"Horizontal alignment: {HORIZONTAL_ALIGNMENTS}",
    )
    vertical: Optional[str] = Field(
        default=None, description=f"Vertical alignment: {VERTICAL_ALIGNMENTS}",
    )
    border_style: Optional[str] = Field(
        default=None, description=f"Border style: {BORDER_STYLES}",
    )
    number_format: Optional[str] = Field(
        default=None, description="Number format string (e.g., '#,##0.00', 'yyyy-mm-dd')",
    )
    wrap_text: Optional[bool] = Field(default=None, description="Enable text wrapping")


class MergeInput(SheetInput):
    range_str: str = Field(..., description="Cell range to merge (e.g., 'A1:C1')")


class SortInput(SheetInput):
    column: str = Field(..., description="Column letter to sort by (e.g., 'A')")
    ascending: bool = Field(default=True, description="Sort ascending if True")
    has_header: bool = Field(default=True, description="Whether data has a header row")


class FilterInput(SheetInput):
    range_str: str = Field(..., description="Range to apply auto-filter (e.g., 'A1:F100')")


class DupesInput(SheetInput):
    columns: Optional[list[str]] = Field(
        default=None, description="Column letters to check for duplicates. All columns if omitted.",
    )


class FormulaInput(SheetInput):
    cell: str = Field(..., description="Cell to place formula (e.g., 'D2')")
    formula: str = Field(..., description="Formula string starting with '=' (e.g., '=SUM(A1:A10)')")

    @field_validator("formula")
    @classmethod
    def _starts_eq(cls, v: str) -> str:
        v = v.strip()
        if not v.startswith("="):
            raise ValueError("Formula must start with '='")
        return v


class FindReplaceInput(SheetInput):
    find: str = Field(..., description="Text to find")
    replace: str = Field(default="", description="Replacement text")
    match_case: bool = Field(default=False)
    match_entire_cell: bool = Field(default=False)


class DescribeInput(SheetInput):
    include_columns: Optional[list[str]] = Field(
        default=None, description="Column letters to describe. All numeric columns if omitted.",
    )


class ExportCSVInput(SheetInput):
    output_path: str = Field(
        ..., description="Destination CSV file path",
    )
    delimiter: str = Field(default=",", description="CSV delimiter character")
    encoding: str = Field(default="utf-8")
    include_index: bool = Field(default=False)


class ImportCSVInput(FilePathInput):
    sheet: str = Field(default="Sheet1", description="Target sheet name")
    csv_path: str = Field(..., description="Source CSV file path")
    delimiter: str = Field(default=",", description="CSV delimiter")
    encoding: str = Field(default="utf-8")


class ExportJSONInput(SheetInput):
    orient: str = Field(
        default="records", description="JSON orientation: 'records', 'index', 'columns', 'values', 'split', 'table'",
    )


class CreateWorkbookInput(BaseModel):
    model_config = ConfigDict(str_strip_whitespace=True)
    filepath: str = Field(
        ..., description="Path for the new Excel file (must end with .xlsx)",
        min_length=1, max_length=4096,
    )
    sheet_name: str = Field(
        default="Sheet1", description="Name for the first sheet",
        min_length=1, max_length=31,
    )
    overwrite: bool = Field(
        default=False, description="Overwrite if file already exists",
    )


# ===================================================================
# TOOLS
# ===================================================================

# ---------- File & Sheet Management ----------

@mcp.tool(
    name="excel_create_workbook",
    annotations={
        "title": "Create Excel Workbook",
        "readOnlyHint": False,
        "destructiveHint": True,
        "idempotentHint": False,
        "openWorldHint": False,
    },
)
def excel_create_workbook(params: CreateWorkbookInput) -> str:
    """Create a new empty Excel workbook (.xlsx).

    Args:
        params (CreateWorkbookInput): filepath, optional sheet_name, overwrite flag.

    Returns:
        str: Success or error message.
    """
    path = _resolve_path(params.filepath)
    if os.path.exists(path) and not params.overwrite:
        return f"Error: File already exists: {path}. Set overwrite=True to replace it."
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = params.sheet_name
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)
    return f"Created workbook: {path} with sheet '{params.sheet_name}'"


@mcp.tool(
    name="excel_list_sheets",
    annotations={
        "title": "List Excel Sheets",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_list_sheets(params: FilePathInput) -> str:
    """List all sheet/tab names in an Excel workbook.

    Args:
        params (FilePathInput): filepath of the Excel file.

    Returns:
        str: JSON list of sheet names or error message.
    """
    try:
        wb = _safe_open(params.filepath)
        sheets = wb.sheetnames
        wb.close()
        return json.dumps(sheets, indent=2, ensure_ascii=False)
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_create_sheet",
    annotations={
        "title": "Create Sheet",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": False,
        "openWorldHint": False,
    },
)
def excel_create_sheet(params: SheetInput) -> str:
    """Add a new sheet to an existing workbook.

    Args:
        params (SheetInput): filepath, sheet name.

    Returns:
        str: Success or error message.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.sheet in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' already exists."
        wb.create_sheet(params.sheet)
        wb.save(path)
        wb.close()
        return f"Created sheet '{params.sheet}' in {path}"
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_delete_sheet",
    annotations={
        "title": "Delete Sheet",
        "readOnlyHint": False,
        "destructiveHint": True,
        "idempotentHint": False,
        "openWorldHint": False,
    },
)
def excel_delete_sheet(params: SheetInput) -> str:
    """Delete a sheet from a workbook.

    Args:
        params (SheetInput): filepath, sheet name.

    Returns:
        str: Success or error message.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found."
        if len(wb.sheetnames) == 1:
            wb.close()
            return "Error: Cannot delete the only sheet in a workbook."
        del wb[params.sheet]
        wb.save(path)
        wb.close()
        return f"Deleted sheet '{params.sheet}' from {path}"
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_rename_sheet",
    annotations={
        "title": "Rename Sheet",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": False,
        "openWorldHint": False,
    },
)
def excel_rename_sheet(params: SheetCopyInput) -> str:
    """Rename a sheet. source_sheet = old name, target_sheet = new name.

    Args:
        params (SheetCopyInput): filepath, source_sheet (old), target_sheet (new).

    Returns:
        str: Success or error message.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.source_sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.source_sheet}' not found."
        if params.target_sheet in wb.sheetnames:
            wb.close()
            return f"Error: Target name '{params.target_sheet}' already exists."
        wb[params.source_sheet].title = params.target_sheet
        wb.save(path)
        wb.close()
        return f"Renamed sheet '{params.source_sheet}' -> '{params.target_sheet}' in {path}"
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_copy_sheet",
    annotations={
        "title": "Copy Sheet",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": False,
        "openWorldHint": False,
    },
)
def excel_copy_sheet(params: SheetCopyInput) -> str:
    """Copy a sheet within the same workbook.

    Args:
        params (SheetCopyInput): filepath, source_sheet, target_sheet (new copy name).

    Returns:
        str: Success or error message.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.source_sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Source sheet '{params.source_sheet}' not found."
        if params.target_sheet in wb.sheetnames:
            wb.close()
            return f"Error: Target name '{params.target_sheet}' already exists."
        source = wb[params.source_sheet]
        target = wb.copy_worksheet(source)
        target.title = params.target_sheet
        wb.save(path)
        wb.close()
        return f"Copied '{params.source_sheet}' -> '{params.target_sheet}' in {path}"
    except Exception as e:
        return f"Error: {e}"


# ---------- Reading ----------

@mcp.tool(
    name="excel_read_sheet",
    annotations={
        "title": "Read Entire Sheet",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_read_sheet(params: SheetInput) -> str:
    """Read all data from a sheet as a Markdown table.

    Args:
        params (SheetInput): filepath, sheet.

    Returns:
        str: Markdown formatted table of all data.
    """
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found. Available: {wb.sheetnames}"
        ws = wb[params.sheet]
        data = _read_sheet_data(ws, with_headers=True)
        wb.close()
        return _format_markdown_table(data, max_rows=100)
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_read_range",
    annotations={
        "title": "Read Cell Range",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_read_range(params: ReadRangeInput) -> str:
    """Read a specific cell range from a sheet and return as Markdown/JSON.

    Args:
        params (ReadRangeInput): filepath, sheet, start_cell, optional end_cell, preview_only.

    Returns:
        str: Markdown table or JSON of the range data.
    """
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found. Available: {wb.sheetnames}"
        ws = wb[params.sheet]

        match = re.match(r"^([A-Z]+)(\d+)$", params.start_cell)
        start_col = column_index_from_string(match.group(1))
        start_row = int(match.group(2))

        if params.end_cell:
            match_e = re.match(r"^([A-Z]+)(\d+)$", params.end_cell)
            end_col = column_index_from_string(match_e.group(1))
            end_row = int(match_e.group(2))
        else:
            end_col = ws.max_column
            end_row = ws.max_row

        data = _read_sheet_data(ws, start_row, end_row, start_col, end_col, with_headers=True)

        wb.close()

        if params.preview_only or len(data["rows"]) <= 20:
            return _format_markdown_table(data, max_rows=50)
        return json.dumps(data, indent=2, ensure_ascii=False, default=str)
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_read_cell",
    annotations={
        "title": "Read Single Cell",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_read_cell(params: CellWriteInput) -> str:
    """Read the value of a single cell.

    Args:
        params (CellWriteInput): filepath, sheet, cell.

    Returns:
        str: JSON with cell reference, value, and data type.
    """
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found."
        ws = wb[params.sheet]
        cell_obj = ws[params.cell]
        result = {
            "cell": params.cell,
            "value": _serialise_value(cell_obj.value),
            "type": type(cell_obj.value).__name__ if cell_obj.value is not None else "None",
            "number_format": cell_obj.number_format,
        }
        wb.close()
        return json.dumps(result, indent=2, ensure_ascii=False, default=str)
    except Exception as e:
        return f"Error: {e}"


# ---------- Writing ----------

@mcp.tool(
    name="excel_write_cell",
    annotations={
        "title": "Write Cell Value",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_write_cell(params: CellWriteInput) -> str:
    """Write a value to a specific cell. Creates/replaces the value.

    Args:
        params (CellWriteInput): filepath, sheet, cell, value.

    Returns:
        str: Success or error message.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found."
        ws = wb[params.sheet]
        ws[params.cell] = params.value
        wb.save(path)
        wb.close()
        return f"Wrote '{params.value}' to {params.cell} in '{params.sheet}'"
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_write_rows",
    annotations={
        "title": "Write Multiple Rows",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_write_rows(params: RowsWriteInput) -> str:
    """Write multiple rows of data starting from a given cell.

    Args:
        params (RowsWriteInput): filepath, sheet, start_cell, data (2D list), header flag.

    Returns:
        str: Success message with count of rows written.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found."
        ws = wb[params.sheet]

        match = re.match(r"^([A-Z]+)(\d+)$", params.start_cell)
        start_col = column_index_from_string(match.group(1))
        start_row = int(match.group(2))

        for r_idx, row in enumerate(params.data):
            for c_idx, val in enumerate(row):
                ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=val)

        wb.save(path)
        wb.close()
        rows_written = len(params.data) - (1 if params.header else 0)
        return f"Wrote {len(params.data)} rows ({rows_written} data rows) to '{params.sheet}' starting at {params.start_cell}"
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_write_tabular",
    annotations={
        "title": "Write Tabular Data (List of Dicts)",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_write_tabular(params: TabularWriteInput) -> str:
    """Write a list of dictionaries as a table to a sheet. Keys become column headers.

    Args:
        params (TabularWriteInput): filepath, sheet, data (list of dicts).

    Returns:
        str: Success message with row/column count.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found."
        ws = wb[params.sheet]

        if not params.data:
            wb.close()
            return "Error: data list is empty."

        headers = list(params.data[0].keys())
        for c_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=c_idx, value=h)

        for r_idx, row in enumerate(params.data, 2):
            for c_idx, h in enumerate(headers, 1):
                ws.cell(row=r_idx, column=c_idx, value=row.get(h))

        wb.save(path)
        wb.close()
        return f"Wrote {len(params.data)} rows x {len(headers)} columns to '{params.sheet}'"
    except Exception as e:
        return f"Error: {e}"


# ---------- Formatting ----------

@mcp.tool(
    name="excel_set_column_width",
    annotations={
        "title": "Set Column Width",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_set_column_width(params: ColWidthInput) -> str:
    """Set the width of one or more columns.

    Args:
        params (ColWidthInput): filepath, sheet, column (letter or range like 'A:C'), width.

    Returns:
        str: Success or error message.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found."
        ws = wb[params.sheet]

        if ":" in params.column:
            start, end = params.column.split(":")
            start_idx = column_index_from_string(start)
            end_idx = column_index_from_string(end)
            for c in range(start_idx, end_idx + 1):
                ws.column_dimensions[get_column_letter(c)].width = params.width
        else:
            ws.column_dimensions[params.column].width = params.width

        wb.save(path)
        wb.close()
        return f"Set column(s) '{params.column}' width to {params.width}"
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_set_row_height",
    annotations={
        "title": "Set Row Height",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_set_row_height(params: RowHeightInput) -> str:
    """Set the height of a specific row.

    Args:
        params (RowHeightInput): filepath, sheet, row number, height.

    Returns:
        str: Success or error message.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found."
        ws = wb[params.sheet]
        ws.row_dimensions[params.row].height = params.height
        wb.save(path)
        wb.close()
        return f"Set row {params.row} height to {params.height}"
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_format_range",
    annotations={
        "title": "Format Cell Range",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_format_range(params: FormatRangeInput) -> str:
    """Apply formatting (font, alignment, fill, border, number format) to a cell range.

    Args:
        params (FormatRangeInput): filepath, sheet, range_str, and any formatting options.

    Returns:
        str: Success message or error.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found."
        ws = wb[params.sheet]

        cells = ws[params.range_str]
        if not isinstance(cells, tuple):
            cells = (cells,)

        applied = []
        for row in cells:
            for cell in row:
                if params.bold is not None or params.font_size is not None or params.font_color is not None:
                    current_font = cell.font
                    cell.font = Font(
                        name=current_font.name,
                        size=params.font_size if params.font_size is not None else current_font.size,
                        bold=params.bold if params.bold is not None else current_font.bold,
                        color=params.font_color if params.font_color else current_font.color,
                    )
                if params.bg_color:
                    cell.fill = PatternFill(start_color=params.bg_color, end_color=params.bg_color, fill_type="solid")
                if params.horizontal or params.vertical or params.wrap_text is not None:
                    current_alignment = cell.alignment or Alignment()
                    cell.alignment = Alignment(
                        horizontal=params.horizontal or current_alignment.horizontal,
                        vertical=params.vertical or current_alignment.vertical,
                        wrap_text=params.wrap_text if params.wrap_text is not None else current_alignment.wrap_text,
                    )
                if params.border_style:
                    cell.border = _build_border(style=params.border_style)
                if params.number_format is not None:
                    cell.number_format = params.number_format

        wb.save(path)
        wb.close()

        applied = []
        if params.bold is not None: applied.append("bold")
        if params.font_size: applied.append(f"font_size={params.font_size}")
        if params.font_color: applied.append(f"font_color={params.font_color}")
        if params.bg_color: applied.append(f"bg_color={params.bg_color}")
        if params.horizontal: applied.append(f"halign={params.horizontal}")
        if params.vertical: applied.append(f"valign={params.vertical}")
        if params.wrap_text is not None: applied.append(f"wrap_text={params.wrap_text}")
        if params.border_style: applied.append(f"border={params.border_style}")
        if params.number_format: applied.append(f"numfmt={params.number_format}")

        return f"Formatted range '{params.range_str}' in '{params.sheet}': {', '.join(applied) if applied else 'no changes'}"
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_merge_cells",
    annotations={
        "title": "Merge Cells",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_merge_cells(params: MergeInput) -> str:
    """Merge a range of cells.

    Args:
        params (MergeInput): filepath, sheet, range_str.

    Returns:
        str: Success or error message.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found."
        ws = wb[params.sheet]
        ws.merge_cells(params.range_str)
        wb.save(path)
        wb.close()
        return f"Merged cells '{params.range_str}' in '{params.sheet}'"
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_unmerge_cells",
    annotations={
        "title": "Unmerge Cells",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_unmerge_cells(params: MergeInput) -> str:
    """Unmerge a previously merged cell range.

    Args:
        params (MergeInput): filepath, sheet, range_str.

    Returns:
        str: Success or error message.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found."
        ws = wb[params.sheet]
        ws.unmerge_cells(params.range_str)
        wb.save(path)
        wb.close()
        return f"Unmerged cells '{params.range_str}' in '{params.sheet}'"
    except Exception as e:
        return f"Error: {e}"


# ---------- Data Manipulation ----------

@mcp.tool(
    name="excel_sort_data",
    annotations={
        "title": "Sort Sheet Data",
        "readOnlyHint": False,
        "destructiveHint": True,
        "idempotentHint": False,
        "openWorldHint": False,
    },
)
def excel_sort_data(params: SortInput) -> str:
    """Sort data in a sheet by a specific column.

    Args:
        params (SortInput): filepath, sheet, column, ascending, has_header.

    Returns:
        str: Success message or error.
    """
    path = _resolve_path(params.filepath)
    try:
        df = pd.read_excel(path, sheet_name=params.sheet)
        col_idx = column_index_from_string(params.column) - 1
        if col_idx >= len(df.columns):
            return f"Error: Column '{params.column}' (index {col_idx}) out of range. Sheet has {len(df.columns)} columns."
        col_name = df.columns[col_idx]
        df = df.sort_values(by=col_name, ascending=params.ascending, ignore_index=True)

        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=params.sheet, index=False)

        return f"Sorted '{params.sheet}' by column '{params.column}' ({'ascending' if params.ascending else 'descending'}). {len(df)} rows."
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_filter_data",
    annotations={
        "title": "Apply Auto-Filter",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_filter_data(params: FilterInput) -> str:
    """Apply auto-filter to a range in a sheet.

    Args:
        params (FilterInput): filepath, sheet, range_str.

    Returns:
        str: Success or error message.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found."
        ws = wb[params.sheet]
        ws.auto_filter.ref = params.range_str
        wb.save(path)
        wb.close()
        return f"Applied auto-filter to '{params.range_str}' in '{params.sheet}'"
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_remove_duplicates",
    annotations={
        "title": "Remove Duplicate Rows",
        "readOnlyHint": False,
        "destructiveHint": True,
        "idempotentHint": False,
        "openWorldHint": False,
    },
)
def excel_remove_duplicates(params: DupesInput) -> str:
    """Remove duplicate rows from a sheet, optionally based on specific columns.

    Args:
        params (DupesInput): filepath, sheet, optional columns list.

    Returns:
        str: Success message with counts.
    """
    path = _resolve_path(params.filepath)
    try:
        df = pd.read_excel(path, sheet_name=params.sheet)
        before = len(df)
        subset = params.columns if params.columns else None
        df = df.drop_duplicates(subset=subset, ignore_index=True)
        after = len(df)
        removed = before - after

        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=params.sheet, index=False)

        cols_msg = f" based on columns {subset}" if subset else ""
        return f"Removed {removed} duplicate rows from '{params.sheet}'{cols_msg}. {after} rows remaining."
    except Exception as e:
        return f"Error: {e}"


# ---------- Formulas ----------

@mcp.tool(
    name="excel_add_formula",
    annotations={
        "title": "Add Excel Formula",
        "readOnlyHint": False,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_add_formula(params: FormulaInput) -> str:
    """Write an Excel formula to a cell.

    Args:
        params (FormulaInput): filepath, sheet, cell, formula.

    Returns:
        str: Success or error message.
    """
    path = _resolve_path(params.filepath)
    try:
        wb = _safe_open(params.filepath)
        if params.sheet not in wb.sheetnames:
            wb.close()
            return f"Error: Sheet '{params.sheet}' not found."
        ws = wb[params.sheet]
        ws[params.cell] = params.formula
        wb.save(path)
        wb.close()
        return f"Added formula '{params.formula}' to {params.cell} in '{params.sheet}'"
    except Exception as e:
        return f"Error: {e}"


# ---------- Search & Replace ----------

@mcp.tool(
    name="excel_find_replace",
    annotations={
        "title": "Find and Replace",
        "readOnlyHint": False,
        "destructiveHint": True,
        "idempotentHint": False,
        "openWorldHint": False,
    },
)
def excel_find_replace(params: FindReplaceInput) -> str:
    """Find and replace text in a sheet.

    Args:
        params (FindReplaceInput): filepath, sheet, find, replace, match_case, match_entire_cell.

    Returns:
        str: Success message with count of replacements.
    """
    path = _resolve_path(params.filepath)
    try:
        df = pd.read_excel(path, sheet_name=params.sheet)

        if params.match_entire_cell:
            mask = df.apply(lambda col: col.astype(str) == params.find)
        else:
            mask = df.apply(lambda col: col.astype(str).str.contains(
                params.find, case=params.match_case, na=False, regex=False,
            ))

        count = mask.sum().sum()
        if count == 0:
            return f"No occurrences of '{params.find}' found in '{params.sheet}'."

        df = df.replace(params.find, params.replace, regex=False)

        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=params.sheet, index=False)

        return f"Replaced {count} occurrence(s) of '{params.find}' with '{params.replace}' in '{params.sheet}'."
    except Exception as e:
        return f"Error: {e}"


# ---------- Analysis ----------

@mcp.tool(
    name="excel_describe_data",
    annotations={
        "title": "Describe/Summarise Data",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_describe_data(params: DescribeInput) -> str:
    """Generate statistical summary of numeric columns in a sheet (count, mean, std, min, quartiles, max).

    Args:
        params (DescribeInput): filepath, sheet, optional include_columns.

    Returns:
        str: Markdown or JSON summary.
    """
    try:
        df = pd.read_excel(params.filepath, sheet_name=params.sheet)

        if params.include_columns:
            cols_in_range = []
            for c in params.include_columns:
                idx = column_index_from_string(c) - 1
                if idx < len(df.columns):
                    cols_in_range.append(df.columns[idx])
            if not cols_in_range:
                return f"Error: None of the specified columns exist in the sheet. Available: {list(df.columns)}"
            df = df[cols_in_range]

        desc = df.describe(include="all").to_string()
        info = io.StringIO()
        df.info(buf=info, verbose=False, show_counts=True)
        info_str = info.getvalue()

        lines = [
            f"# Sheet: '{params.sheet}'",
            f"Rows: {len(df)}, Columns: {len(df.columns)}",
            "",
            "## Column Info",
            f"```\n{info_str}\n```",
            "",
            "## Statistical Summary",
            f"```\n{desc}\n```",
        ]
        return "\n".join(lines)
    except Exception as e:
        return f"Error: {e}"


# ---------- Import / Export ----------

@mcp.tool(
    name="excel_to_csv",
    annotations={
        "title": "Export Sheet to CSV",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_to_csv(params: ExportCSVInput) -> str:
    """Export a sheet from an Excel file to a CSV file.

    Args:
        params (ExportCSVInput): filepath, sheet, output_path, delimiter, encoding, include_index.

    Returns:
        str: Success message or error.
    """
    try:
        df = pd.read_excel(params.filepath, sheet_name=params.sheet)
        df.to_csv(params.output_path, sep=params.delimiter, encoding=params.encoding, index=params.include_index)
        return f"Exported '{params.sheet}' to CSV: {params.output_path} ({len(df)} rows x {len(df.columns)} cols)"
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_csv_to_sheet",
    annotations={
        "title": "Import CSV to Sheet",
        "readOnlyHint": False,
        "destructiveHint": True,
        "idempotentHint": False,
        "openWorldHint": False,
    },
)
def excel_csv_to_sheet(params: ImportCSVInput) -> str:
    """Import a CSV file into a sheet, replacing existing content.

    Args:
        params (ImportCSVInput): filepath, sheet, csv_path, delimiter, encoding.

    Returns:
        str: Success message or error.
    """
    path = _resolve_path(params.filepath)
    try:
        if not os.path.exists(params.csv_path):
            return f"Error: CSV file not found: {params.csv_path}"

        df = pd.read_csv(params.csv_path, sep=params.delimiter, encoding=params.encoding)

        if os.path.exists(path):
            with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df.to_excel(writer, sheet_name=params.sheet, index=False)
        else:
            df.to_excel(path, sheet_name=params.sheet, index=False)

        return f"Imported {params.csv_path} -> '{params.sheet}' in {path} ({len(df)} rows x {len(df.columns)} cols)"
    except Exception as e:
        return f"Error: {e}"


@mcp.tool(
    name="excel_to_json",
    annotations={
        "title": "Export Sheet to JSON",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
def excel_to_json(params: ExportJSONInput) -> str:
    """Export a sheet's data as JSON.

    Args:
        params (ExportJSONInput): filepath, sheet, orient.

    Returns:
        str: JSON string of the sheet data.
    """
    try:
        df = pd.read_excel(params.filepath, sheet_name=params.sheet)
        return df.to_json(orient=params.orient, indent=2, force_ascii=False, default_handler=str)
    except Exception as e:
        return f"Error: {e}"


# ===================================================================
# Entry point
# ===================================================================

if __name__ == "__main__":
    mcp.run()
